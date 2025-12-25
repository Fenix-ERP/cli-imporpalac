import base64
import io
import logging
from io import BytesIO

import xlsxwriter
from openpyxl import load_workbook

from odoo import _, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class AccountHomologate(models.Model):
    _inherit = "account.homologate"

    file = fields.Binary()
    file_name = fields.Char()

    def action_import(self):
        if not self.file:
            raise UserError(_("You must load an archive to be imported."))

        try:
            # Decodificar el archivo base64
            data = base64.b64decode(self.file)
            xlsx_file = BytesIO(data)
            workbook = load_workbook(xlsx_file, data_only=True)
            sheet = workbook.active

            # Recorremos desde la segunda fila (saltamos cabecera)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                supplier_code, product_ref = row[0], row[1]

                if not supplier_code or not product_ref:
                    continue

                product = self.env["product.product"].search(
                    [("default_code", "=", str(product_ref))], limit=1
                )

                if not product:
                    continue

                line = self.lines_prod.filtered(
                    lambda line: line.cod_supplier == str(supplier_code)
                )
                for line_item in line:
                    line_item.product_id = product.id

            self.file = False
            self.file_name = False

        except Exception as e:
            raise UserError(_("Error processing the file: %s") % str(e)) from None

        return {
            "type": "ir.actions.act_window",
            "name": "Homologate",
            "res_model": "account.homologate",
            "view_mode": "form",
            "res_id": self.id,
            "target": "new",
            "context": self.env.context,
        }

    def action_export_template(self):
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})

        header_format = workbook.add_format(
            {
                "bold": True,
                "text_wrap": True,
                "valign": "vcenter",
                "align": "center",
                "border": 1,
                "bg_color": "#D9E1F2",
            }
        )
        example_format = workbook.add_format(
            {"font_color": "#0070C0", "valign": "vcenter", "align": "left", "border": 1}
        )
        headers = [
            _("Code Supplier"),
            _("Description Supplier"),
            _("Product ID"),
            _("Description"),
        ]

        sheet = workbook.add_worksheet("Plantilla Contacto")
        sheet.set_column(0, len(headers) - 1, 25)
        for col, header in enumerate(headers):
            sheet.write(0, col, header, header_format)
        column = 0
        row = 1
        for line in self.lines_prod:
            sheet.write(row, column, line.cod_supplier, example_format)
            sheet.write(row, column + 1, line.description_supplier, example_format)
            sheet.write(row, column + 2, line.product_id.display_name, example_format)
            sheet.write(row, column + 3, line._description, example_format)
            column += 1
            row += 1

        workbook.close()
        file_data = base64.b64encode(output.getvalue())
        output.close()

        attachment = self.env["ir.attachment"].create(
            {
                "name": "contact_template.xlsx",
                "type": "binary",
                "datas": file_data,
                "res_model": "account.homologate",
                "res_id": self.id,
                "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
        )

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }
