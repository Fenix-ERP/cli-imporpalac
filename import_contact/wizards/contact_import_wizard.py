import base64
import io
import re
from io import BytesIO

import requests
import xlsxwriter
from openpyxl import load_workbook
from PIL import Image

from odoo import _, fields, models
from odoo.exceptions import UserError, ValidationError


class ImportContactWizard(models.TransientModel):
    _name = "import.contact.wizard"
    _description = "Import Contact Wizard"

    file = fields.Binary(string="Archivo", required=False)
    types = fields.Selection(
        [
            ("supplier", "Proveedor"),
            ("customer", "Cliente"),
            ("both", "Cliente y Proveedor"),
        ],
        string="Tipo",
        required=True,
    )

    def is_valid_image_base64(self, b64_string):
        try:
            decoded = base64.b64decode(b64_string, validate=True)
            Image.open(BytesIO(decoded)).verify()
            return True
        except Exception:
            return False

    def is_valid_base64(self, string):
        try:
            cleaned = re.sub(r"[^A-Za-z0-9+/=]", "", str(string or ""))
            padding = len(cleaned) % 4
            if padding:
                cleaned += "=" * (4 - padding)
            base64.b64decode(cleaned)
            return True
        except Exception:
            return False

    def get_image_from_drive(self, shared_url):
        import logging

        _logger = logging.getLogger(__name__)
        try:
            session = requests.Session()
            response = session.get(shared_url, stream=True, timeout=15)

            if "Content-Disposition" not in response.headers:
                file_id_match = re.search(r"/d/([a-zA-Z0-9_-]+)", shared_url)
                if not file_id_match:
                    _logger.warning(f"[Drive] Invalid shared link format: {shared_url}")
                    return None
                file_id = file_id_match.group(1)

                confirm_token = None
                for key, value in response.cookies.items():
                    if key.startswith("download_warning"):
                        confirm_token = value
                        break

                download_url = (
                    f"https://drive.google.com/uc?export=download&id={file_id}"
                )
                if confirm_token:
                    download_url += f"&confirm={confirm_token}"

                response = session.get(download_url, stream=True, timeout=15)

            if response.status_code != 200:
                _logger.warning(
                    f"[Drive] Failed to fetch image. Status: {response.status_code}"
                )
                return None

            content_type = response.headers.get("Content-Type", "")
            if "image" not in content_type.lower():
                _logger.warning(f"[Drive] Not an image: Content-Type: {content_type}")
                return None

            img = Image.open(BytesIO(response.content)).convert("RGB")
            buffer = BytesIO()
            img.save(buffer, format="PNG")
            return base64.b64encode(buffer.getvalue()).decode("utf-8")

        except Exception as e:
            _logger.warning(f"[Drive] Error downloading image from shared link: {e}")
            return None

    def get_value(self, row, index):
        try:
            return str(row[index]).strip() if row[index] is not None else ""
        except IndexError:
            return ""

    def action_test(self):
        if not self.file:
            raise UserError(_("Debe subir un archivo."))

        file_data = base64.b64decode(self.file)
        file_format = self._get_file_format(file_data)

        errores = []
        actualizables = 0
        creables = 0

        identificaciones = set()
        ciudades = set()
        provincias = set()
        etiquetas = set()
        cuentas_pagar = set()
        cuentas_cobrar = set()
        plazos_pago = set()

        if self.types == "supplier":
            CAMPOS_OBLIGATORIOS = {
                0: "Es Empresa",
                1: "Nombre",
                3: "Tipo de Identificación",
                4: "Identificación",
                14: "Cuenta por Pagar",
                15: "Cuenta por Cobrar",
                16: "Plazo de Pago",
            }
        elif self.types == "customer":
            CAMPOS_OBLIGATORIOS = {
                0: "Es Empresa",
                1: "Nombre",
                3: "Tipo de Identificación",
                4: "Identificación",
                14: "Cuenta por Pagar",
                15: "Cuenta por Cobrar",
            }

        # Primera pasada para recolectar datos
        data_iter = self._parse_file(file_data, file_format)
        rows_cache = []  # Guarda temporalmente filas ya recorridas

        for i, row in enumerate(data_iter, start=2):
            rows_cache.append(row)
            if not any(cell for cell in row if str(cell).strip()):
                continue
            errores_fila = []

            for idx, label in CAMPOS_OBLIGATORIOS.items():
                valor = self.get_value(row, idx)
                if not valor:
                    errores_fila.append(label)

            if errores_fila:
                errores.append(
                    f"Fila {i}: Campos obligatorios vacíos: {', '.join(errores_fila)}."
                )
                continue

            identificacion = self.get_value(row, 4)
            provincia = self.get_value(row, 5)
            ciudad = self.get_value(row, 6)
            cuenta_por_pagar = self.get_value(row, 14)
            cuenta_por_cobrar = self.get_value(row, 15)
            plazo_de_pago = self.get_value(row, 16)
            etiquetas_raw = self.get_value(row, 13)

            identificaciones.add(identificacion)
            ciudades.add(ciudad)
            provincias.add(provincia)
            cuentas_pagar.add(cuenta_por_pagar)
            cuentas_cobrar.add(cuenta_por_cobrar)
            plazos_pago.add(plazo_de_pago)

            for tag in etiquetas_raw.split(","):
                etiquetas.add(tag.strip())

        # Validaciones de existencia
        partners = self.env["res.partner"].search(
            [("vat", "in", list(identificaciones))]
        )
        identificaciones_existentes = set(partners.mapped("vat"))

        provincias_existentes = set(
            self.env["res.country.state"]
            .search([("name", "in", list(provincias))])
            .mapped("name")
        )
        cuentas_existentes = set(
            self.env["account.account"]
            .search([("code", "in", list(cuentas_pagar))])
            .mapped("code")
        )
        cuentas_existentes.update(
            set(
                self.env["account.account"]
                .search([("code", "in", list(cuentas_cobrar))])
                .mapped("code")
            )
        )
        plazos_existentes = set(
            self.env["account.payment.term"]
            .search([("name", "in", list(plazos_pago))])
            .mapped("name")
        )
        etiquetas_existentes = set(
            self.env["res.partner.category"]
            .search([("name", "in", list(etiquetas))])
            .mapped("name")
        )

        # Segunda pasada sobre las filas almacenadas
        for i, row in enumerate(rows_cache, start=2):
            identificacion = self.get_value(row, 4)
            provincia = self.get_value(row, 5)
            ciudad = self.get_value(row, 6)
            cuenta_por_pagar = self.get_value(row, 14)
            cuenta_por_cobrar = self.get_value(row, 15)
            plazo_de_pago = self.get_value(row, 16)
            etiquetas_raw = self.get_value(row, 13)
            image_input = self.get_value(row, 18)

            if provincia and provincia not in provincias_existentes:
                errores.append(f"Fila {i}: Provincia '{provincia}' no encontrada.")
            if cuenta_por_pagar and cuenta_por_pagar not in cuentas_existentes:
                errores.append(f"Fila {i}: Cuenta '{cuenta_por_pagar}' no encontrada.")
            if cuenta_por_cobrar and cuenta_por_cobrar not in cuentas_existentes:
                errores.append(f"Fila {i}: Cuenta '{cuenta_por_cobrar}' no encontrada.")
            if plazo_de_pago and plazo_de_pago not in plazos_existentes:
                errores.append(
                    f"Fila {i}: Plazo de pago '{plazo_de_pago}' no encontrado."
                )

            for tag in etiquetas_raw.split(","):
                if tag.strip() and tag.strip() not in etiquetas_existentes:
                    errores.append(f"Fila {i}: Etiqueta '{tag.strip()}' no encontrada.")

            if identificacion in identificaciones_existentes:
                actualizables += 1
            else:
                creables += 1

            if image_input:
                is_valid = False
                if self.is_valid_base64(image_input):
                    is_valid = True
                elif (
                    isinstance(image_input, str)
                    and "drive.google.com/file/d/" in image_input
                ):
                    is_valid = True
                elif self.get_image_from_drive(image_input):
                    is_valid = True

                if not is_valid:
                    errores.append(
                        f"Fila {i}: Formato de imagen inválido o enlace de Drive inaccesible."
                    )

        if errores:
            raise ValidationError("\n".join(errores))
        else:
            mensaje = f"Archivo válido.\n\n✅ Contactos a crear: {creables}\n🔁 Contactos a actualizar: {actualizables}"
            return {
                "type": "ir.actions.client",
                "tag": "display_notification",
                "params": {
                    "title": "Simulación de importación de proveedores",
                    "message": mensaje,
                    "type": "success",
                    "sticky": False,
                },
            }

    def action_export_template(self):
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})

        header_format = workbook.add_format(
            {
                "bold": True,
                "text_wrap": True,
                "valign": "vcenter",
                "align": "center",
                "border": 1,
                "bg_color": "#D9E1F2",
            }
        )
        example_format = workbook.add_format(
            {"font_color": "#0070C0", "valign": "vcenter", "align": "left", "border": 1}
        )
        header_notes_format = workbook.add_format(
            {
                "bold": True,
                "text_wrap": True,
                "valign": "vcenter",
                "align": "center",
                "border": 1,
                "bg_color": "#4CAF50",
                "color": "white",
            }
        )
        notes_format = workbook.add_format(
            {
                "text_wrap": True,
                "valign": "top",
                "align": "left",
                "border": 1,
                "bg_color": "#F9F9F9",
            }
        )

        headers = [
            "Es Empresa",
            "Es Cliente",
            "Es Proveedor",
            "Nombre",
            "Nombre Comercial",
            "Tipo de Identificación",
            "Identificación",
            "Provincia",
            "Ciudad",
            "Calle (Dirección 1)",
            "Referencia (Dirección 2)",
            "Teléfono",
            "Celular",
            "Correo Electrónico",
            "Sitio Web",
            "Etiquetas",
            "Cuenta por Cobrar",
            "Cuenta por Pagar",
            "Plazo de Pago",
            "Lista de Precios",
            "Documentos Electrónicos",
            "Imagen",
        ]

        example_data = [
            "True",
            "True",
            "False",
            "Ejemplo S.A.",
            "Comercial Ejemplo",
            "CEDULA",
            "0912345678",
            "Pichincha",
            "Quito",
            "Av. Amazonas",
            "Edificio ABC",
            "+59322345678",
            "+593998765432",
            "info@ejemplo.com",
            "https://ejemplo.com",
            "Importador,Cliente VIP",
            "ANTICIPOS EMPLEADO",
            "ANTICIPOS CLIENTE",
            "Pago Inmediato",
            "Lista de Precios Estándar",
            "True",
            "https://drive.google.com/file/d/1ABCDEF23456789GHIJKL/view?usp=sharing",
        ]

        notes = [
            'Es Empresa: "True" si el contacto es una empresa, "False" en caso contrario.',
            'Es Cliente: "True" si el contacto es un cliente.',
            'Es Proveedor: "True" si el contacto es un proveedor.',
            "Nombre: Obligatorio. Nombre principal del contacto o empresa.",
            "Nombre Comercial: Opcional. Nombre alternativo.",
            'Tipo de Identificación: Debe ser "CEDULA", "RUC" o "PASAPORTE".',
            "Identificación: Número de 10 a 13 dígitos. Solo se permiten números.",
            "Provincia: Debe coincidir con un nombre de provincia registrado.",
            "Ciudad: Obligatorio. Ciudad del contacto.",
            "Calle (Dirección 1): Calle o avenida principal.",
            "Referencia (Dirección 2): Referencia adicional (ej. edificio).",
            'Teléfono: Opcional. Usa formatos como "+593...", "09..." o fijos como "02 2345678".',
            'Celular: Debe comenzar con "+593" o "09".',
            "Correo Electrónico: Obligatorio. Debe ser válido.",
            "Sitio Web: Opcional. Debe ser una URL válida.",
            "Etiquetas: Opcional. Nombres separados por comas (deben existir en el sistema).",
            "Cuenta por Cobrar: Debe coincidir con un nombre de cuenta existente.",
            "Cuenta por Pagar: Debe coincidir con un nombre de cuenta existente.",
            "Plazo de Pago: Debe coincidir con un nombre de término de pago existente.",
            "Lista de Precios: Debe coincidir con un nombre de lista de precios existente.",
            'Documentos Electrónicos: "True" para habilitar la facturación electrónica.',
            'Imagen: Debes proporcionar un enlace compartido a una imagen almacenada en Google Drive. El archivo debe estar compartido con la opción "Cualquiera con el enlace - Lector" y el enlace debe ser copiado usando la opción "Copiar enlace" desde Google Drive. Ejemplo: https://drive.google.com/file/d/FILE_ID/view?usp=sharing',
        ]

        sheet = workbook.add_worksheet("Plantilla Contacto")
        sheet.set_column(0, len(headers) - 1, 25)
        for col, header in enumerate(headers):
            sheet.write(0, col, header, header_format)
        for col, example in enumerate(example_data):
            sheet.write(1, col, example, example_format)

        sheet_notes = workbook.add_worksheet("Observaciones")
        sheet_notes.set_column("A:A", 90)
        sheet_notes.write("A1", "Observaciones", header_notes_format)
        for row, note in enumerate(notes, start=1):
            sheet_notes.write(row, 0, note, notes_format)

        workbook.close()
        file_data = base64.b64encode(output.getvalue())
        output.close()

        attachment = self.env["ir.attachment"].create(
            {
                "name": "contact_template.xlsx",
                "type": "binary",
                "datas": file_data,
                "res_model": "import.contact.wizard",
                "res_id": self.id,
                "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
        )

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }

    def action_import(self):
        if not self.file:
            raise UserError(_("Debe subir un archivo."))

        file_data = base64.b64decode(self.file)
        file_format = self._get_file_format(file_data)
        data = self._parse_file(file_data, file_format)

        for i, row in enumerate(data, start=2):
            if not any(cell for cell in row if str(cell).strip()):
                continue

            nombre = self.get_value(row, 1)
            id_type_clean = self.get_value(row, 3).upper()
            tipo_identificacion = {
                "CEDULA": "cedula",
                "RUC": "ruc",
                "PASAPORTE": "pasaporte",
            }.get(id_type_clean)
            identificacion = self.get_value(row, 4)
            provincia_nombre = self.get_value(row, 5)
            ciudad = self.get_value(row, 6)
            calle = self.get_value(row, 7)
            calle2 = self.get_value(row, 8)
            telefono = self.get_value(row, 9)
            celular = self.get_value(row, 10)
            email = self.get_value(row, 11)
            sitio_web = self.get_value(row, 12)
            etiquetas_raw = self.get_value(row, 13)
            cuenta_por_pagar_nombre = self.get_value(row, 14)
            cuenta_por_cobrar_nombre = self.get_value(row, 15)
            plazo_nombre = self.get_value(row, 16)
            internacional = self.get_value(row, 17)
            image_input = self.get_value(row, 18)
            # Buscar o crear registros relacionados
            provincia = self.env["res.country.state"].search(
                [("name", "=", provincia_nombre)], limit=1
            )
            cuenta_por_pagar = self.env["account.account"].search(
                [("code", "=", cuenta_por_pagar_nombre)], limit=1
            )
            cuenta_por_cobrar = self.env["account.account"].search(
                [("code", "=", cuenta_por_cobrar_nombre)], limit=1
            )
            plazo = self.env["account.payment.term"].search(
                [("name", "=", plazo_nombre)], limit=1
            )

            etiquetas_ids = []
            for tag in etiquetas_raw.split(","):
                tag = tag.strip()
                if not tag:
                    continue
                etiqueta = self.env["res.partner.category"].search(
                    [("name", "=", tag)], limit=1
                )
                if etiqueta:
                    etiquetas_ids.append(etiqueta.id)

            # Buscar proveedor
            partner = self.env["res.partner"].search(
                [("vat", "=", identificacion)], limit=1
            )
            supplier_rank = 0
            customer_rank = 0

            if self.types == "supplier":
                supplier_rank = 1
            elif self.types == "customer":
                customer_rank = 1
            elif self.types == "both":
                supplier_rank = 1
                customer_rank = 1

            if self.types == "supplier":

                vals = {
                    "name": nombre,
                    "identifier": identificacion,
                    "type_identifier": tipo_identificacion,
                    "city": ciudad,
                    "state_id": provincia.id if provincia else False,
                    "street": calle,
                    "street2": calle2,
                    "phone": telefono,
                    "mobile": celular,
                    "email": email,
                    "website": sitio_web,
                    "supplier_rank": supplier_rank,
                    "customer_rank": customer_rank,
                    "property_account_payable_id": cuenta_por_pagar.id
                    if cuenta_por_pagar
                    else False,
                    "property_account_receivable_id": cuenta_por_cobrar.id
                    if cuenta_por_cobrar
                    else False,
                    "property_supplier_payment_term_id": plazo.id if plazo else False,
                    "is_international": internacional if internacional else False,
                    "category_id": [(6, 0, etiquetas_ids)],
                }

            elif self.types == "customer":
                limite_credito = self.get_value(row, 19)
                valor_limite = self.get_value(row, 20)
                lista_precios = self.get_value(row, 21)
                metodo_pago = self.get_value(row, 22)
                vendedor = self.get_value(row, 23)
                price_list = self.env["product.pricelist"].search(
                    [("name", "=", lista_precios)], limit=1
                )
                payment_method = self.env["payment.type"].search(
                    [("name", "=", metodo_pago)], limit=1
                )
                vendor = self.env["res.partner"].search(
                    [("name", "=", vendedor)], limit=1
                )
                vals = {
                    "name": nombre,
                    "identifier": identificacion,
                    "type_identifier": tipo_identificacion,
                    "city": ciudad,
                    "state_id": provincia.id if provincia else False,
                    "street": calle,
                    "street2": calle2,
                    "phone": telefono,
                    "mobile": celular,
                    "email": email,
                    "website": sitio_web,
                    "supplier_rank": supplier_rank,
                    "customer_rank": customer_rank,
                    "property_account_payable_id": cuenta_por_pagar.id
                    if cuenta_por_pagar
                    else False,
                    "property_account_receivable_id": cuenta_por_cobrar.id
                    if cuenta_por_cobrar
                    else False,
                    "property_payment_term_id": plazo.id if plazo else False,
                    "is_international": internacional if internacional else False,
                    "category_id": [(6, 0, etiquetas_ids)],
                    "use_partner_credit_limit": limite_credito
                    if limite_credito
                    else False,
                    "credit_limit": valor_limite if valor_limite else False,
                    "property_product_pricelist": price_list.id
                    if price_list
                    else False,
                    "payment_method": payment_method.id if payment_method else False,
                    "user_id": vendor.id if vendor else False,
                }

            # Imagen
            if image_input:
                if self.is_valid_base64(image_input):
                    vals["image_1920"] = image_input
                elif "drive.google.com/file/d/" in image_input:
                    image_data = self.get_image_from_drive(image_input)
                    if image_data:
                        vals["image_1920"] = image_data

            # Crear o actualizar
            if partner:
                partner.write(vals)
            else:
                self.env["res.partner"].create(vals)

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": "Importación exitosa",
                "message": "Los contactos han sido importados correctamente.",
                "type": "success",
                "sticky": False,
            },
        }

    def _get_file_format(self, file_data):
        """Detecta el tipo de archivo según su contenido"""
        if file_data[:4] == b"PK\x03\x04":  # Este es el encabezado de archivos XLSX
            return "xlsx"
        elif file_data[:3] == b"\xef\xbb\xbf":  # BOM de archivos CSV (UTF-8 con BOM)
            return "csv"

        return ""

    def _parse_file(self, file_data, file_format):
        if file_format != "xlsx":
            raise UserError(_("Formato de archivo no compatible. Solo se admite .xlsx"))

        wb = load_workbook(filename=BytesIO(file_data), read_only=True, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(
                cell not in [None, "", " "] and str(cell).strip() != "" for cell in row
            ):
                continue
            yield row
