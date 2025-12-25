import base64
import io
from io import BytesIO

from openpyxl import load_workbook

from odoo import _, fields, models
from odoo.exceptions import UserError, ValidationError


class ImportProductWizard(models.TransientModel):
    _name = "import.product.wizard"
    _description = "Import Product Wizard"

    file = fields.Binary(string="Archivo", required=False)

    def action_test(self):
        if not self.file:
            raise UserError(_("Debe subir un archivo."))

        file_data = base64.b64decode(self.file)
        file_format = self._get_file_format(file_data)

        errores = []
        actualizables = 0
        creables = 0

        categorias_dict = {
            c.display_name.strip(): c.id
            for c in self.env["product.category"].search([])
        }
        productos_dict = {
            p.default_code.strip(): p
            for p in self.env["product.template"].search([])
            if p.default_code
        }

        # Primera pasada para recolectar datos
        data_iter = self._parse_file(file_data, file_format)
        rows_cache = []  # Guarda temporalmente filas ya recorridas
        UOM = self.env["uom.uom"]
        uoms_dict = {u.name.lower(): u.id for u in UOM.search([])}

        for i, row in enumerate(data_iter, start=2):
            rows_cache.append(row)
            if not any(cell for cell in row if str(cell).strip()):
                continue
            try:
                if not row or len(row) < 16:
                    continue

                codigo = str(row[0]).strip()
                categoria_nombre = row[16].strip()
                unidad_venta = row[14]
                unidad_compra = row[15]

                uom_id = uoms_dict.get(unidad_venta.lower())
                uom_po_id = uoms_dict.get(unidad_compra.lower())

                if not uom_id:
                    errores.append(
                        f"Fila {i+1}: Unidad de venta '{unidad_venta}' no encontrada"
                    )
                    continue

                if not uom_po_id:
                    errores.append(
                        f"Fila {i+1}: Unidad de compra '{unidad_compra}' no encontrada"
                    )
                    continue

                if not codigo:
                    errores.append(f"Fila {i+1}: Código vacío")
                    continue

                if categoria_nombre not in categorias_dict:
                    errores.append(
                        f"Fila {i+1}: Categoría '{categoria_nombre}' no encontrada"
                    )
                    continue

                if codigo in productos_dict:
                    actualizables += 1
                else:
                    creables += 1

            except Exception as e:
                errores.append(f"Fila {i+1}: Error inesperado: {str(e)}")

        if errores:
            raise ValidationError("\n".join(errores))
        else:
            mensaje = f"Archivo válido.\n\n✅ Líneas que se actualizarían: {actualizables}\n🔁 Líneas que se crearían: {creables}"
            return {
                "type": "ir.actions.client",
                "tag": "display_notification",
                "params": {
                    "title": "Simulación de importación",
                    "message": mensaje,
                    "type": "success",
                    "sticky": False,
                },
            }

    def _get_file_format(self, file_data):
        """Detecta el tipo de archivo según su contenido"""
        if file_data[:4] == b"PK\x03\x04":  # Este es el encabezado de archivos XLSX
            return "xlsx"
        elif file_data[:3] == b"\xef\xbb\xbf":  # BOM de archivos CSV (UTF-8 con BOM)
            return "csv"

        return ""

    def action_export_template(self):

        import xlsxwriter

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})

        # Hoja principal
        worksheet = workbook.add_worksheet("Plantilla Lista de Precios")

        header_format = workbook.add_format(
            {
                "bold": True,
                "text_wrap": True,
                "valign": "vcenter",
                "align": "center",
                "border": 1,
                "bg_color": "#D9E1F2",
            }
        )

        sample_format_text = workbook.add_format(
            {
                "text_wrap": True,
                "valign": "top",
                "align": "left",
                "border": 1,
                "bg_color": "#FFF2CC",
            }
        )

        sample_format_number = workbook.add_format(
            {
                "valign": "top",
                "align": "right",
                "border": 1,
                "bg_color": "#FFF2CC",
                "num_format": "#,##0.00",
            }
        )

        headers = [
            "Nombre",
            "Codigo Producto",
            "Producto",
            "Cantidad Minima",
            "Precio",
            "Fecha Inicio",
            "Fecha Finalizacion",
        ]

        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_format)
        worksheet.set_column(0, len(headers) - 1, 30)

        # Fila de ejemplo
        example_row = [
            "Mayor",
            "GDBT0539A",
            "Capot Aluminio Peugeot 3008 2021/",
            0,
            555,
            "",
            "",
        ]
        for col_num, value in enumerate(example_row):
            if col_num in [3, 4]:  # Cantidad Minima, Precio
                worksheet.write_number(1, col_num, value, sample_format_number)
            else:
                worksheet.write(1, col_num, value, sample_format_text)

        # Hoja de observaciones
        worksheet2 = workbook.add_worksheet("Observaciones")
        header_observations_format = workbook.add_format(
            {
                "bold": True,
                "text_wrap": True,
                "valign": "vcenter",
                "align": "center",
                "border": 1,
                "bg_color": "#4CAF50",
                "color": "white",
            }
        )

        observations_format = workbook.add_format(
            {
                "bold": False,
                "text_wrap": True,
                "valign": "top",
                "align": "left",
                "border": 1,
                "bg_color": "#F9F9F9",
            }
        )

        worksheet2.write("A1", "Observaciones", header_observations_format)

        observaciones = [
            "Nombre: Nombre exacto de la lista de precios (case sensitive).",
            "Código Producto: Código interno del producto (default_code).",
            "Producto: Nombre del producto (referencial, no se utiliza para vincular).",
            "Cantidad Minima: Número entero o decimal. No debe dejarse vacío.",
            "Precio: Precio fijo a aplicar (decimal).",
            "Fecha Inicio: Formato YYYY-MM-DD.",
            "Fecha Finalizacion: Formato YYYY-MM-DD.",
        ]

        for row, obs in enumerate(observaciones, start=1):
            worksheet2.write(row, 0, obs, observations_format)

        worksheet2.set_column("A:A", 90)

        workbook.close()
        file_data = base64.b64encode(output.getvalue())
        output.close()

        attachment = self.env["ir.attachment"].create(
            {
                "name": "plantilla_lista_precios.xlsx",
                "type": "binary",
                "datas": file_data,
                "res_model": self._name,
                "res_id": self.id,
                "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
        )

        return {
            "type": "ir.actions.act_url",
            "url": "/web/content/%s?download=true" % attachment.id,
            "target": "self",
        }

    def action_import(self):
        if not self.file:
            raise UserError(_("Debe subir un archivo."))

        file_data = base64.b64decode(self.file)
        file_format = self._get_file_format(file_data)

        errores = []
        creados = 0
        actualizados = 0
        nuevos_vals = []

        Product = self.env["product.template"]
        Category = self.env["product.category"]
        UOM = self.env["uom.uom"]

        categorias_dict = {c.display_name.strip(): c.id for c in Category.search([])}
        productos_dict = {
            p.default_code.strip(): p for p in Product.search([]) if p.default_code
        }
        uoms_dict = {u.name.lower(): u.id for u in UOM.search([])}

        data = self._parse_file(file_data, file_format)

        def procesar_lote(vals_list):
            nonlocal creados
            try:
                Product.create(vals_list)
                creados += len(vals_list)
            except Exception as e:
                errores.append(f"Error en lote: {repr(e)}")
                # Fallback: uno por uno con savepoint
                for vals in vals_list:
                    try:
                        with self.env.cr.savepoint():
                            Product.create(vals)
                            creados += 1
                    except Exception as e2:
                        errores.append(
                            f"Producto con código {vals.get('default_code')}: Error al crear - {str(e2)}"
                        )

        for i, row in enumerate(data, start=2):
            if not any(str(cell).strip() for cell in row):
                continue
            try:
                if len(row) < 25:
                    errores.append(f"Fila {i}: No tiene todas las columnas necesarias.")
                    continue

                row = [
                    str(cell).strip() if isinstance(cell, str) else cell for cell in row
                ]

                codigo, nombre, product_article = row[0], row[1], row[2]
                product_brand, product_model, product_year = (
                    row[3] or "",
                    row[4] or "",
                    row[5] or "",
                )
                product_origin = row[6] or ""
                adicional = row[7:13]
                product_side = row[13] or ""
                unidad_venta = row[14]
                unidad_compra = row[15]
                categoria = row[16]
                costo = float(row[17] or 0.0)

                vender = str(row[18]).lower() in ["1", "true", "sí", "si"]
                comprar = str(row[19]).lower() in ["1", "true", "sí", "si"]
                importar = str(row[20]).lower() in ["1", "true", "sí", "si"]
                tipo_producto = row[21]

                categoria_importacion = row[22] or ""
                politica_control = row[23]
                politica_facturacion = row[24]

                categoria_id = categorias_dict.get(categoria)
                if not categoria_id:
                    errores.append(f"Fila {i}: Categoría '{categoria}' no encontrada.")
                    continue

                tipo_map = {
                    "Producto almacenable": "product",
                    "Consumible": "consu",
                    "Servicio": "service",
                }
                tipo_producto = tipo_map.get(tipo_producto, "product")

                control_map = {
                    "Sobre cantidades recibidas": "receive",
                    "Sobre cantidades pedidas": "purchase",
                }
                purchase_method = control_map.get(politica_control, "receive")

                factura_map = {
                    "Cantidades entregadas": "delivery",
                    "Cantidad ordenada": "order",
                }
                invoice_policy = factura_map.get(politica_facturacion, "delivery")

                uom_id = uoms_dict.get(unidad_venta.lower())
                uom_po_id = uoms_dict.get(unidad_compra.lower())

                vals = {
                    "default_code": codigo,
                    "product_article": product_article,
                    "name": nombre,
                    "categ_id": categoria_id,
                    "standard_price": costo,
                    "sale_ok": vender,
                    "purchase_ok": comprar,
                    "type": tipo_producto,
                    "uom_id": uom_id,
                    "uom_po_id": uom_po_id,
                    "product_brand": product_brand,
                    "product_model": product_model,
                    "product_year": product_year,
                    "product_origin": product_origin,
                    "additional_1": adicional[0],
                    "additional_2": adicional[1],
                    "additional_3": adicional[2],
                    "additional_4": adicional[3],
                    "additional_5": adicional[4],
                    "additional_6": adicional[5],
                    "product_side": product_side,
                    "imported_ok": importar,
                    "import_category_id": categoria_importacion,
                    "purchase_method": purchase_method,
                    "invoice_policy": invoice_policy,
                }

                producto = productos_dict.get(codigo)
                if producto:
                    producto.write(vals)
                    actualizados += 1
                else:
                    nuevos_vals.append(vals)
                    if len(nuevos_vals) >= 100:
                        procesar_lote(nuevos_vals)
                        nuevos_vals.clear()

            except Exception as e:
                errores.append(f"Fila {i}: Error inesperado: {str(e)}")

        # Lote final pendiente
        if nuevos_vals:
            procesar_lote(nuevos_vals)

        mensaje = f"Importación completada.\n\nProductos actualizados: {actualizados}\nProductos creados: {creados}"
        if errores:
            mensaje += f"\n\nErrores detectados:\n- " + "\n- ".join(errores[:10])

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": "Resultado de importación",
                "message": mensaje,
                "type": "warning" if errores else "success",
                "sticky": False,
            },
        }

    def _parse_file(self, file_data, file_format):
        if file_format != "xlsx":
            raise UserError(_("Formato de archivo no compatible. Solo se admite .xlsx"))

        wb = load_workbook(filename=BytesIO(file_data), read_only=True, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(
                cell not in [None, "", " "] and str(cell).strip() != "" for cell in row
            ):
                continue
            yield row
