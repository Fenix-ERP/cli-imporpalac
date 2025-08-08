import base64
import logging
from io import BytesIO

from openpyxl import load_workbook

from odoo import _, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class AccountHomologate(models.Model):
    _inherit = "account.homologate"

    file = fields.Binary()
    file_name = fields.Char()

    def action_import(self):
        if not self.file:
            raise UserError(_("Debe cargar un archivo para importar."))

        try:
            # Decodificar el archivo base64
            data = base64.b64decode(self.file)
            xlsx_file = BytesIO(data)
            workbook = load_workbook(xlsx_file, data_only=True)
            sheet = workbook.active

            for row in enumerate(
                sheet.iter_rows(min_row=2), start=2
            ):  # desde la segunda fila
                supplier_code = row[0].value
                product_ref = row[1].value

                if not supplier_code or not product_ref:
                    continue

                product = self.env["product.product"].search(
                    [("default_code", "=", str(product_ref))], limit=1
                )

                if not product:
                    continue

                line = self.lines_prod.filtered(
                    lambda line: line.cod_supplier == str(supplier_code)
                )
                for line_item in line:
                    line_item.product_id = product.id

            self.file = False
            self.file_name = False
        except Exception as e:
    raise UserError(_("Error procesando el archivo: %s") % str(e)) from None

        return {
            "type": "ir.actions.act_window",
            "name": "Homologate",
            "res_model": "account.homologate",
            "view_mode": "form",
            "res_id": self.id,
            "target": "new",
            "context": self.env.context,
        }
