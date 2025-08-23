import base64
from io import BytesIO

from openpyxl import load_workbook

from odoo import _, fields, models
from odoo.exceptions import UserError


class ImportTransferProductWizard(models.TransientModel):
    _name = "import.transfer.product.wizard"
    _description = "Import Transfer Product Wizard"

    file = fields.Binary(required=True)

    def _get_file_format(self, file_data):
        """Detect file type"""
        if file_data[:4] == b"PK\x03\x04":  # XLSX
            return "xlsx"
        return ""

    def action_import(self):
        if not self.file:
            raise UserError(_("You must upload a file."))

        file_data = base64.b64decode(self.file)
        file_format = self._get_file_format(file_data)

        if file_format != "xlsx":
            raise UserError(_("File format not supported. Only .xlsx is supported."))

        # obtenemos el picking activo
        picking = self.env["stock.picking"].browse(self._context.get("active_id"))
        if not picking:
            raise UserError(_("No picking found to import lines."))

        Product = self.env["product.product"]

        productos_dict = {
            p.default_code.strip(): p for p in Product.search([]) if p.default_code
        }

        errores = []
        creados = 0

        wb = load_workbook(filename=BytesIO(file_data), read_only=True, data_only=True)
        sheet = wb.active

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue

            try:
                default_code = str(row[0]).strip()
                demanda = float(row[1] or 0.0)

                product = productos_dict.get(default_code)
                if not product:
                    errores.append(f"Row {i}: Product with reference '{default_code}' not found.")
                    continue

                if demanda <= 0:
                    errores.append(f"Row {i}: Invalid demand ({demanda}).")
                    continue

                self.env["stock.move"].create({
                    "name": product.display_name,
                    "product_id": product.id,
                    "product_uom": product.uom_id.id,
                    "product_uom_qty": demanda,
                    "picking_id": picking.id,
                    "location_id": picking.location_id.id,
                    "location_dest_id": picking.location_dest_id.id,
                })
                creados += 1

            except Exception as e:
                errores.append(_(f"Row {i}: Unexpected error: {str(e)}"))

        mensaje = _("Import completed.\n\nMovements created: %(count)s", count=creados)
        if errores:
            mensaje += f"\n\nErrors detected:\n- " + "\n- ".join(errores[:10])

        return {
            "effect": {
                "fadeout": "slow",
                "message": mensaje,
                "type": "rainbow_man",
            }
        }
